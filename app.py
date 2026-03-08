from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, flash
)
import sqlite3
import os
import re
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook
from urllib.parse import quote

app = Flask(__name__)
app.secret_key = "salma-violations-secret-key"

# =========================
# إعدادات النظام
# =========================
DB_NAME = "student_violations.db"
ADMIN_USERNAME = "سلمى"
ADMIN_PASSWORD = "مخالفات@95"
AUTO_IMPORT_FILE = os.path.join("uploads", "students.xlsx")

SYSTEM_NAME = "مدرسة سلمى بنت قيس للتعليم الأساسي (5–12)"
SYSTEM_SUBTITLE = "نظام إدارة مخالفات قواعد الانضباط السلوكي"
SYSTEM_ARTICLE = "لائحة شؤون الطلاب – المادة (39)"

LEVEL_LABELS = {
    "A": "نصح",
    "B": "تنبيه",
    "C": "إنذار",
    "D": "فصل"
}

LEVEL_EMOJIS = {
    "A": "🟢",
    "B": "🟡",
    "C": "🟠",
    "D": "🔴"
}

LEVEL_CLASSES = {
    "A": "level-a",
    "B": "level-b",
    "C": "level-c",
    "D": "level-d"
}


# =========================
# أدوات قاعدة البيانات
# =========================
def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def query_one(sql, params=()):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(sql, params)
    row = cur.fetchone()
    conn.close()
    return row


def query_all(sql, params=()):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return rows


def execute_db(sql, params=()):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(sql, params)
    conn.commit()
    last_id = cur.lastrowid
    conn.close()
    return last_id


def execute_many(sql, data):
    conn = get_db()
    cur = conn.cursor()
    cur.executemany(sql, data)
    conn.commit()
    conn.close()


# =========================
# إنشاء الجداول
# =========================
def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_name TEXT NOT NULL,
            class_name TEXT,
            father_name TEXT,
            father_phone TEXT,
            mother_name TEXT,
            mother_phone TEXT,
            area TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS violations_catalog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            base_level TEXT NOT NULL CHECK(base_level IN ('A','B','C','D')),
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS escalation_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_level TEXT NOT NULL,
            target_level TEXT NOT NULL,
            threshold INTEGER NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS violation_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            violation_id INTEGER NOT NULL,
            base_level TEXT NOT NULL,
            result_level TEXT NOT NULL,
            note TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(student_id) REFERENCES students(id),
            FOREIGN KEY(violation_id) REFERENCES violations_catalog(id)
        )
    """)

    conn.commit()
    conn.close()

    seed_default_settings()
    seed_default_violations()
    auto_import_students_if_needed()


# =========================
# البيانات الافتراضية
# =========================
def seed_default_settings():
    count = query_one("SELECT COUNT(*) AS c FROM escalation_settings")["c"]
    if count == 0:
        rows = [
            ("A", "C", 3),
            ("A", "D", 5),
            ("B", "C", 2),
            ("B", "D", 4),
            ("C", "D", 3),
        ]
        execute_many("""
            INSERT INTO escalation_settings (source_level, target_level, threshold)
            VALUES (?, ?, ?)
        """, rows)


def seed_default_violations():
    count = query_one("SELECT COUNT(*) AS c FROM violations_catalog")["c"]
    if count > 0:
        return

    defaults = [
        # A - نصح
        ("الإخلال بنظام الطابور أو الحصص الدراسية أو الأنشطة المدرسية", "A"),
        ("عدم الالتزام بالزي المدرسي", "A"),
        ("العبث بمرافق المدرسة ووسائل النقل المدرسية", "A"),
        ("الإساءة بالقول إلى أحد زملائها", "A"),
        ("عدم المحافظة على النظافة أو المظهر الشخصي", "A"),
        ("عدم الالتزام بإحضار الكتب والدفاتر والأدوات المدرسية والملابس الرياضية", "A"),
        ("عدم مراعاة الاحترام الواجب في التعامل مع معلمي وفنيي وإداريي المدرسة وزوارها", "A"),
        ("تناول المأكولات أو المشروبات في غير الوقت المخصص أو مضغ العلكة", "A"),
        ("النوم أثناء الحصص أو الأنشطة المدرسية", "A"),
        ("عدم المحافظة على نظافة الفصل وغيره من مرافق المدرسة", "A"),
        ("الإهمال في أداء الواجبات أو الأنشطة المدرسية", "A"),
        ("عدم الإنصات لتوجيهات المعلم", "A"),
        ("التسبب في الإزعاج بالقرب من الصفوف الدراسية", "A"),
        ("عدم الالتزام بضوابط استخدام وسائل النقل المدرسية", "A"),
        ("القيام بأي سلوك يخالف الآداب والنظام العام", "A"),
        ("أخرى - نصح", "A"),

        # B - تنبيه
        ("تكرار أحد السلوكيات المنصوص عليها في بند النصح", "B"),
        ("تزوير أحد المحررات المدرسية", "B"),
        ("تزوير توقيع ولي الأمر", "B"),

        # C - إنذار
        ("تكرار أحد السلوكيات المنصوص عليها في بند النصح للمرة الثالثة", "C"),
        ("تكرار السلوكيات المنصوص عليها في بندي تزوير المحررات أو توقيع ولي الأمر للمرة الثانية", "C"),
        ("الشجار وتهديد زملائها", "C"),
        ("الاعتداء بألفاظ نابية على أحد زملائها", "C"),
        ("الاستيلاء على المتعلقات الشخصية لزملائها", "C"),
        ("أخرى - إنذار", "C"),

        # D - فصل
        ("تكرار أحد السلوكيات حسب قواعد التصعيد المعتمدة", "D"),
        ("إحضار الأجهزة السمعية والبصرية كالهواتف النقالة والكاميرات والمسجلات وغيرها في غير الأغراض التعليمية", "D"),
        ("حيازة أو تداول أو استعمال مواد ضارة كالكبريت والألعاب النارية والأدوات الحادة وغيرها", "D"),
        ("الاعتداء على زملائها وإيذائهم بدنيًا", "D"),
        ("الكتابة على جدران المدرسة", "D"),
        ("الهروب من المدرسة", "D"),
        ("أخرى - فصل", "D"),
    ]

    execute_many("""
        INSERT INTO violations_catalog (title, base_level)
        VALUES (?, ?)
    """, defaults)


# =========================
# استيراد تلقائي أول تشغيل من Excel
# =========================
def normalize_phone(phone):
    if phone is None:
        return ""

    phone = str(phone).strip()

    if not phone or phone.lower() == "none":
        return ""

    # إزالة أي رموز أو مسافات
    phone = re.sub(r"[^\d]", "", phone)

    if not phone:
        return ""

    # إذا يبدأ بـ 00 نحوله لدولي بدون 00
    if phone.startswith("00"):
        phone = phone[2:]

    # إذا الرقم العماني محلي 8 أرقام نضيف 968
    if len(phone) == 8:
        phone = "968" + phone

    # إذا مكتوب 968 ورائه 8 أرقام فهو صحيح
    # إذا فيه + أو فراغات قد أزلناها فوق

    return phone


def auto_import_students_if_needed():
    student_count = query_one("SELECT COUNT(*) AS c FROM students")["c"]
    if student_count > 0:
        return

    if not os.path.exists(AUTO_IMPORT_FILE):
        return

    try:
        wb = load_workbook(AUTO_IMPORT_FILE)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}

        def get_value(row, col_name):
            pos = idx.get(col_name)
            if pos is None:
                return ""
            val = row[pos].value
            return str(val).strip() if val is not None else ""

        rows_to_insert = []

        for row in ws.iter_rows(min_row=2):
            student_name = get_value(row, "اسم الطالبه الثلاثي مع القبيلة")
            if not student_name:
                continue

            class_name = get_value(row, "الصف")
            father_name = get_value(row, "الاسم الثلاثي ولي الامر")
            father_phone = normalize_phone(get_value(row, "رقم هاتف ولي الامر"))
            mother_name = get_value(row, "الاسم الثلاثي للأم")
            mother_phone = normalize_phone(get_value(row, "رقم هاتف الأم"))
            area = get_value(row, "المنطقة السكنيه")

            rows_to_insert.append((
                student_name, class_name, father_name, father_phone,
                mother_name, mother_phone, area
            ))

        if rows_to_insert:
            execute_many("""
                INSERT INTO students (
                    student_name, class_name, father_name, father_phone,
                    mother_name, mother_phone, area
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, rows_to_insert)

    except Exception as e:
        print("Auto import error:", e)


# =========================
# التحقق من تسجيل الدخول
# =========================
def login_required():
    return session.get("logged_in") is True


@app.before_request
def require_login():
    allowed = ["login", "static"]
    if request.endpoint in allowed:
        return
    if request.endpoint is None:
        return
    if not login_required():
        return redirect(url_for("login"))


# =========================
# منطق التصعيد
# =========================
def calculate_result_level(base_level, total_count_after_save):
    rules = query_all("""
        SELECT source_level, target_level, threshold
        FROM escalation_settings
        WHERE source_level = ?
        ORDER BY threshold DESC
    """, (base_level,))

    result = base_level
    for rule in rules:
        if total_count_after_save >= rule["threshold"]:
            result = rule["target_level"]
            break
    return result


def get_same_violation_count(student_id, violation_id):
    row = query_one("""
        SELECT COUNT(*) AS c
        FROM violation_records
        WHERE student_id = ? AND violation_id = ?
    """, (student_id, violation_id))
    return row["c"] if row else 0


def level_badge(level):
    return f"{LEVEL_EMOJIS.get(level, '')} {LEVEL_LABELS.get(level, level)}"


def build_whatsapp_link(phone, message):
    phone = normalize_phone(phone)
    if not phone:
        return ""

    encoded_message = quote(message)

    # فتح مباشر على واتساب ويب
    return f"https://web.whatsapp.com/send?phone={phone}&text={encoded_message}"


def build_message(student, violation_title, result_level, note=""):
    result_text = LEVEL_LABELS.get(result_level, result_level)
    msg = (
        f"السلام عليكم ورحمة الله وبركاته\n"
        f"نفيدكم بأنه تم تسجيل مخالفة سلوكية على الطالبة: {student['student_name']}\n"
        f"الصف: {student['class_name'] or '-'}\n"
        f"نوع المخالفة: {violation_title}\n"
        f"الإجراء المتخذ: {result_text}\n"
    )
    if note:
        msg += f"ملاحظة: {note}\n"
    msg += f"المدرسة: {SYSTEM_NAME}"
    return msg


# =========================
# التقارير
# =========================
def get_dashboard_data():
    students = query_all("""
        SELECT *
        FROM students
        ORDER BY student_name COLLATE NOCASE
    """)

    violations = query_all("""
        SELECT *
        FROM violations_catalog
        ORDER BY
            CASE base_level
                WHEN 'A' THEN 1
                WHEN 'B' THEN 2
                WHEN 'C' THEN 3
                WHEN 'D' THEN 4
            END,
            title
    """)

    settings = query_all("""
        SELECT *
        FROM escalation_settings
        ORDER BY source_level, threshold
    """)

    records = query_all("""
        SELECT vr.id, vr.created_at, vr.note, vr.base_level, vr.result_level,
               s.student_name, s.class_name,
               v.title AS violation_title
        FROM violation_records vr
        JOIN students s ON s.id = vr.student_id
        JOIN violations_catalog v ON v.id = vr.violation_id
        ORDER BY vr.id DESC
    """)

    summary = {
        "total_students": query_one("SELECT COUNT(*) AS c FROM students")["c"],
        "total_records": query_one("SELECT COUNT(*) AS c FROM violation_records")["c"],
        "a_count": query_one("SELECT COUNT(*) AS c FROM violation_records WHERE result_level='A'")["c"],
        "b_count": query_one("SELECT COUNT(*) AS c FROM violation_records WHERE result_level='B'")["c"],
        "c_count": query_one("SELECT COUNT(*) AS c FROM violation_records WHERE result_level='C'")["c"],
        "d_count": query_one("SELECT COUNT(*) AS c FROM violation_records WHERE result_level='D'")["c"],
    }

    by_violation = query_all("""
        SELECT v.title, COUNT(*) AS c
        FROM violation_records vr
        JOIN violations_catalog v ON v.id = vr.violation_id
        GROUP BY vr.violation_id
        ORDER BY c DESC
    """)

    by_level = query_all("""
        SELECT result_level, COUNT(*) AS c
        FROM violation_records
        GROUP BY result_level
        ORDER BY result_level
    """)

    monthly = query_all("""
        SELECT strftime('%Y-%m', created_at) AS month, COUNT(*) AS c
        FROM violation_records
        GROUP BY month
        ORDER BY month
    """)

    top_class = query_one("""
        SELECT s.class_name, COUNT(*) AS c
        FROM violation_records vr
        JOIN students s ON s.id = vr.student_id
        GROUP BY s.class_name
        ORDER BY c DESC
        LIMIT 1
    """)

    low_class = query_one("""
        SELECT s.class_name, COUNT(*) AS c
        FROM violation_records vr
        JOIN students s ON s.id = vr.student_id
        GROUP BY s.class_name
        ORDER BY c ASC
        LIMIT 1
    """)

    top_student = query_one("""
        SELECT s.student_name, COUNT(*) AS c
        FROM violation_records vr
        JOIN students s ON s.id = vr.student_id
        GROUP BY s.student_name
        ORDER BY c DESC
        LIMIT 1
    """)

    top_violation = query_one("""
        SELECT v.title, COUNT(*) AS c
        FROM violation_records vr
        JOIN violations_catalog v ON v.id = vr.violation_id
        GROUP BY v.title
        ORDER BY c DESC
        LIMIT 1
    """)

    return {
        "students": students,
        "violations": violations,
        "settings": settings,
        "records": records,
        "summary": summary,
        "by_violation": by_violation,
        "by_level": by_level,
        "monthly": monthly,
        "top_class": top_class,
        "low_class": low_class,
        "top_student": top_student,
        "top_violation": top_violation,
        "level_labels": LEVEL_LABELS,
        "level_classes": LEVEL_CLASSES,
        "system_name": SYSTEM_NAME,
        "system_subtitle": SYSTEM_SUBTITLE,
        "system_article": SYSTEM_ARTICLE
    }


# =========================
# صفحات
# =========================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("dashboard"))

        flash("بيانات الدخول غير صحيحة", "error")

    return render_template(
        "login.html",
        system_name=SYSTEM_NAME,
        system_subtitle=SYSTEM_SUBTITLE,
        system_article=SYSTEM_ARTICLE
    )


@app.route("/dashboard")
def dashboard():
    data = get_dashboard_data()
    return render_template("dashboard.html", **data)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# =========================
# API - الطالبات
# =========================
@app.route("/api/student/<int:student_id>")
def api_student(student_id):
    student = query_one("SELECT * FROM students WHERE id = ?", (student_id,))
    if not student:
        return jsonify({"ok": False, "message": "الطالبة غير موجودة"}), 404
    return jsonify({"ok": True, "student": dict(student)})


@app.route("/student/save", methods=["POST"])
def student_save():
    student_id = request.form.get("student_id", "").strip()
    student_name = request.form.get("student_name", "").strip()
    class_name = request.form.get("class_name", "").strip()
    father_name = request.form.get("father_name", "").strip()
    father_phone = normalize_phone(request.form.get("father_phone", "").strip())
    mother_name = request.form.get("mother_name", "").strip()
    mother_phone = normalize_phone(request.form.get("mother_phone", "").strip())
    area = request.form.get("area", "").strip()

    if not student_name:
        flash("اسم الطالبة مطلوب", "error")
        return redirect(url_for("dashboard"))

    if student_id:
        execute_db("""
            UPDATE students
            SET student_name=?, class_name=?, father_name=?, father_phone=?,
                mother_name=?, mother_phone=?, area=?
            WHERE id=?
        """, (
            student_name, class_name, father_name, father_phone,
            mother_name, mother_phone, area, student_id
        ))
        flash("تم تعديل بيانات الطالبة", "success")
    else:
        execute_db("""
            INSERT INTO students (
                student_name, class_name, father_name, father_phone,
                mother_name, mother_phone, area
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            student_name, class_name, father_name, father_phone,
            mother_name, mother_phone, area
        ))
        flash("تمت إضافة الطالبة", "success")

    return redirect(url_for("dashboard") + "#students-section")


@app.route("/student/delete/<int:student_id>", methods=["POST"])
def student_delete(student_id):
    used = query_one("""
        SELECT COUNT(*) AS c FROM violation_records WHERE student_id = ?
    """, (student_id,))["c"]

    if used > 0:
        flash("لا يمكن حذف الطالبة لأن لها سجلات مخالفات", "error")
    else:
        execute_db("DELETE FROM students WHERE id = ?", (student_id,))
        flash("تم حذف الطالبة", "success")

    return redirect(url_for("dashboard") + "#students-section")


@app.route("/students/import", methods=["POST"])
def students_import():
    file = request.files.get("excel_file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("يرجى رفع ملف Excel بصيغة xlsx", "error")
        return redirect(url_for("dashboard") + "#students-section")

    try:
        wb = load_workbook(file)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}

        required_name = "اسم الطالبه الثلاثي مع القبيلة"
        if required_name not in idx:
            flash("ملف Excel لا يحتوي العمود الأساسي: اسم الطالبه الثلاثي مع القبيلة", "error")
            return redirect(url_for("dashboard") + "#students-section")

        def get_value(row, col_name):
            pos = idx.get(col_name)
            if pos is None:
                return ""
            val = row[pos].value
            return str(val).strip() if val is not None else ""

        inserted = 0
        for row in ws.iter_rows(min_row=2):
            student_name = get_value(row, "اسم الطالبه الثلاثي مع القبيلة")
            if not student_name:
                continue

            class_name = get_value(row, "الصف")
            father_name = get_value(row, "الاسم الثلاثي ولي الامر")
            father_phone = normalize_phone(get_value(row, "رقم هاتف ولي الامر"))
            mother_name = get_value(row, "الاسم الثلاثي للأم")
            mother_phone = normalize_phone(get_value(row, "رقم هاتف الأم"))
            area = get_value(row, "المنطقة السكنيه")

            exists = query_one("""
                SELECT id FROM students
                WHERE student_name = ? AND class_name = ?
            """, (student_name, class_name))

            if exists:
                continue

            execute_db("""
                INSERT INTO students (
                    student_name, class_name, father_name, father_phone,
                    mother_name, mother_phone, area
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                student_name, class_name, father_name, father_phone,
                mother_name, mother_phone, area
            ))
            inserted += 1

        flash(f"تم استيراد {inserted} طالبة بنجاح", "success")

    except Exception as e:
        flash(f"حدث خطأ أثناء استيراد الملف: {e}", "error")

    return redirect(url_for("dashboard") + "#students-section")


# =========================
# API - المخالفات
# =========================
@app.route("/api/preview-action", methods=["POST"])
def preview_action():
    data = request.get_json()
    student_id = data.get("student_id")
    violation_id = data.get("violation_id")
    note = (data.get("note") or "").strip()

    if not student_id or not violation_id:
        return jsonify({"ok": False, "message": "اختاري الطالبة والمخالفة"}), 400

    student = query_one("SELECT * FROM students WHERE id=?", (student_id,))
    violation = query_one("""
        SELECT * FROM violations_catalog
        WHERE id=? AND is_active=1
    """, (violation_id,))

    if not student or not violation:
        return jsonify({"ok": False, "message": "بيانات غير صحيحة"}), 400

    previous_count = get_same_violation_count(student_id, violation_id)
    total_after = previous_count + 1
    result_level = calculate_result_level(violation["base_level"], total_after)

    message = build_message(student, violation["title"], result_level, note)
    father_link = build_whatsapp_link(student["father_phone"], message)
    mother_link = build_whatsapp_link(student["mother_phone"], message)

    return jsonify({
        "ok": True,
        "student": dict(student),
        "violation": dict(violation),
        "previous_count": previous_count,
        "count_after_save": total_after,
        "result_level": result_level,
        "result_label": LEVEL_LABELS[result_level],
        "result_badge": level_badge(result_level),
        "father_link": father_link,
        "mother_link": mother_link
    })


@app.route("/record/save", methods=["POST"])
def record_save():
    student_id = request.form.get("student_id", "").strip()
    violation_id = request.form.get("violation_id", "").strip()
    note = request.form.get("note", "").strip()

    if not student_id or not violation_id:
        flash("اختاري الطالبة والمخالفة أولًا", "error")
        return redirect(url_for("dashboard") + "#records-section")

    student = query_one("SELECT * FROM students WHERE id=?", (student_id,))
    violation = query_one("""
        SELECT * FROM violations_catalog
        WHERE id=? AND is_active=1
    """, (violation_id,))

    if not student or not violation:
        flash("بيانات الطالبة أو المخالفة غير صحيحة", "error")
        return redirect(url_for("dashboard") + "#records-section")

    previous_count = get_same_violation_count(student["id"], violation["id"])
    total_after = previous_count + 1
    result_level = calculate_result_level(violation["base_level"], total_after)

    execute_db("""
        INSERT INTO violation_records (
            student_id, violation_id, base_level, result_level, note, created_at
        ) VALUES (?, ?, ?, ?, ?, ?)
    """, (
        student["id"], violation["id"], violation["base_level"],
        result_level, note, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))

    flash(f"تم حفظ السجل بنجاح - الإجراء الناتج: {LEVEL_LABELS[result_level]}", "success")
    return redirect(url_for("dashboard") + "#records-section")


@app.route("/record/delete/<int:record_id>", methods=["POST"])
def record_delete(record_id):
    execute_db("DELETE FROM violation_records WHERE id=?", (record_id,))
    flash("تم حذف السجل", "success")
    return redirect(url_for("dashboard") + "#reports-section")


@app.route("/catalog/save", methods=["POST"])
def catalog_save():
    violation_id = request.form.get("violation_id", "").strip()
    title = request.form.get("title", "").strip()
    base_level = request.form.get("base_level", "").strip()

    if not title or base_level not in ("A", "B", "C", "D"):
        flash("بيانات البند غير مكتملة", "error")
        return redirect(url_for("dashboard") + "#catalog-section")

    if violation_id:
        execute_db("""
            UPDATE violations_catalog
            SET title=?, base_level=?
            WHERE id=?
        """, (title, base_level, violation_id))
        flash("تم تعديل البند", "success")
    else:
        execute_db("""
            INSERT INTO violations_catalog (title, base_level)
            VALUES (?, ?)
        """, (title, base_level))
        flash("تمت إضافة البند", "success")

    return redirect(url_for("dashboard") + "#catalog-section")


@app.route("/catalog/toggle/<int:violation_id>", methods=["POST"])
def catalog_toggle(violation_id):
    row = query_one("SELECT is_active FROM violations_catalog WHERE id=?", (violation_id,))
    if row:
        new_status = 0 if row["is_active"] == 1 else 1
        execute_db("UPDATE violations_catalog SET is_active=? WHERE id=?", (new_status, violation_id))
        flash("تم تحديث حالة البند", "success")
    return redirect(url_for("dashboard") + "#catalog-section")


@app.route("/catalog/delete/<int:violation_id>", methods=["POST"])
def catalog_delete(violation_id):
    used = query_one("""
        SELECT COUNT(*) AS c FROM violation_records WHERE violation_id=?
    """, (violation_id,))["c"]

    if used > 0:
        flash("لا يمكن حذف البند لأنه مرتبط بسجلات", "error")
    else:
        execute_db("DELETE FROM violations_catalog WHERE id=?", (violation_id,))
        flash("تم حذف البند", "success")

    return redirect(url_for("dashboard") + "#catalog-section")


@app.route("/settings/save", methods=["POST"])
def settings_save():
    try:
        a_to_c = int(request.form.get("a_to_c", "3"))
        a_to_d = int(request.form.get("a_to_d", "5"))
        b_to_c = int(request.form.get("b_to_c", "2"))
        b_to_d = int(request.form.get("b_to_d", "4"))
        c_to_d = int(request.form.get("c_to_d", "3"))

        execute_db("DELETE FROM escalation_settings")

        rows = [
            ("A", "C", a_to_c),
            ("A", "D", a_to_d),
            ("B", "C", b_to_c),
            ("B", "D", b_to_d),
            ("C", "D", c_to_d),
        ]

        execute_many("""
            INSERT INTO escalation_settings (source_level, target_level, threshold)
            VALUES (?, ?, ?)
        """, rows)

        flash("تم حفظ إعدادات التصعيد", "success")
    except Exception as e:
        flash(f"خطأ في إعدادات التصعيد: {e}", "error")

    return redirect(url_for("dashboard") + "#catalog-section")


# =========================
# تنزيل Excel
# =========================
@app.route("/export/excel")
def export_excel():
    filter_class = request.args.get("class_name", "").strip()
    filter_month = request.args.get("month", "").strip()
    filter_year = request.args.get("year", "").strip()

    sql = """
        SELECT vr.created_at, s.student_name, s.class_name,
               v.title AS violation_title,
               vr.result_level, vr.note
        FROM violation_records vr
        JOIN students s ON s.id = vr.student_id
        JOIN violations_catalog v ON v.id = vr.violation_id
        WHERE 1=1
    """
    params = []

    if filter_class:
        sql += " AND s.class_name = ?"
        params.append(filter_class)

    if filter_month:
        sql += " AND strftime('%Y-%m', vr.created_at) = ?"
        params.append(filter_month)

    if filter_year:
        sql += " AND strftime('%Y', vr.created_at) = ?"
        params.append(filter_year)

    sql += " ORDER BY vr.id DESC"

    rows = query_all(sql, tuple(params))

    wb = Workbook()
    ws = wb.active
    ws.title = "المخالفات"

    headers = ["التاريخ والوقت", "اسم الطالبة", "الصف", "نوع المخالفة", "المستوى الناتج", "ملاحظة"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row["created_at"],
            row["student_name"],
            row["class_name"],
            row["violation_title"],
            LEVEL_LABELS.get(row["result_level"], row["result_level"]),
            row["note"] or ""
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="violations_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================
# تشغيل التطبيق
# =========================
if __name__ == "__main__":
    os.makedirs("uploads", exist_ok=True)
    init_db()
    app.run(debug=True)